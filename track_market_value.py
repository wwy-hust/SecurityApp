# -*- coding:utf-8 -*-
import akshare as ak
import pandas as pd
import numpy as np
import math
import os, sys
import traceback
import time

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabel, DataLabelList

from StockInfoLibrary.StockInfoClass import StockInfoProxy
from StockInfoLibrary.AkShareDataHelper import *
from StockInfoLibrary.TypeDefine import CodeType, TypeInPortfolio


####################### Settings #######################
CHART = False
####################### Settings #######################

Stock = ['603087', '00700', 'SY']
COLOFFSET = {
	'代码&名称': 0,
	'摘要': 1,
	'年份': 2,
	'净利润': 3,
	'增长率': 4,
	'Q1净利': 5,
	'Q1增长': 6,
	'Q2净利': 7,
	'Q2增长': 8,
	'Q3净利': 9,
	'Q3增长': 10,
	'Q4净利': 11,
	'Q4增长': 12,
	'净利Sum': 13,
	'低估估值': 14,
	'合理估值': 15,
	'高估估值': 16,
	'当前市值': 17,
}


def is_contains_chinese(strs):
	for _char in strs:
		if '\u4e00' <= _char <= '\u9fa5':
			return True
	return False


def containInValidChar(strs):
	for c in ('.', '*', ',', '`', '!', '@', '#', '$', '%', '^', '&', '(', ')'):
		if c in strs:
			return True
	return False


def isCodeValid(code):
	code = str(code)
	isValid = code != 'nan' and not is_contains_chinese(code) and not containInValidChar(code)
	# print('checking code', code, isValid)
	return isValid


def getStockCodeListFromExcel(excelPath, sheetName, codeTitle):
	if os.path.exists(excelPath):
		sheet = pd.read_excel(excelPath, sheet_name=sheetName)
		codes = sheet[codeTitle]
		validCode = list(filter(lambda c: isCodeValid(c), codes))
		codes = [str(c) for c in validCode]
		return codes


def trackMarketValue(excelPath):
	global Stock

	wb = openpyxl.load_workbook(excelPath)
	for sheetName in wb.sheetnames:#('高确定性', ):
		ws = wb[sheetName]
		clearAllChart(ws)
		retDict = getStockCodeListFromExcel(excelPath, sheetName, '代码&名称')
		if retDict:
			Stock = retDict
		print(sheetName, ":", Stock)
		ret = {}
		ROW_CNT = 7
		for idx, stock in enumerate(Stock):
			realstockCode = "a." + stock
			stockInfo = StockInfoProxy(realstockCode)
			stockInfo.fetchCodeData() #(True)
			ret[stock] = stockInfo
			updateToExcel(ws, 2 + (idx * ROW_CNT), 1, 2 + (idx + 1) * ROW_CNT, 18, stockInfo)
		for _, cell in ws._cells.items():
			cell.alignment = Alignment(horizontal='center', vertical='center')

	wb.save(excelPath)


def updateToExcel(worksheet, startRow, startCol, endRow, endCol, stockInfo):
	ws = worksheet
	rowCnt = endRow - startRow
	colCnt = endCol - startCol

	print(stockInfo.name, "mk_v:", stockInfo.market_value, "pe_ttm:", stockInfo.pe_ttm)

	for row in range(rowCnt):
		s1valid, s2valid, s3valid, s4valid = False, False, False, False
		s1profit, s2profit, s3profit, s4profit = 0, 0, 0, 0
		curRowYear = 0
		for col in range(colCnt):
			if col == COLOFFSET['年份'] and row in (1, 2, 3, 4, 5):
				# 年份
				if ws.cell(row=startRow + row, column=startCol + col).value in ("", None):
					ws.cell(row=startRow + row, column=startCol + col).value = str(time.localtime().tm_year + row - 2)
				curRowYear = ws.cell(row=startRow + row, column=startCol + col).value
				continue
			elif col == COLOFFSET['增长率'] and row in (2, 3, 4, 5):
				# 逐年增长率
				if ws.cell(row=startRow + row, column=startCol + col).value in ("", None):
					ws.cell(row=startRow + row, column=startCol + col).value = "20%"
				continue
			elif col == COLOFFSET['净利润'] and row == 1:
				# 去年净利润
				if ws.cell(row=startRow + row, column=startCol + col).value in ("", None):
					ws.cell(row=startRow + row, column=startCol + col).value = "1"
				continue
			elif col in (COLOFFSET['低估估值'], COLOFFSET['合理估值'], COLOFFSET['高估估值']) and row == 0:
				if ws.cell(row=startRow + row, column=startCol + col).value in ("", None):
					ws.cell(row=startRow + row, column=startCol + col).value = "25"
				continue
			elif col in (COLOFFSET['Q1净利'], COLOFFSET['Q2净利'], COLOFFSET['Q3净利'], COLOFFSET['Q4净利']):
				if stockInfo.profit is None:
					continue
				reportDates = reportDateList(int(curRowYear))
				if col == COLOFFSET['Q1净利']:
					row_s1 = stockInfo.profit.loc[lambda df:df['REPORT_DATE'] =='%s 00:00:00' % reportDates[0],:]
					if len(row_s1):
						s1profit = float(row_s1['NETPROFIT'].values[0])
						ws.cell(row=startRow + row, column=startCol + col).value = "%02.02f" % (s1profit / 100000000)
						s1valid = True
				elif col == COLOFFSET['Q2净利']:
					row_s2 = stockInfo.profit.loc[lambda df:df['REPORT_DATE'] =='%s 00:00:00' % reportDates[1],:]
					if len(row_s2):
						s2profit = row_s2['NETPROFIT'].values[0]
						ws.cell(row=startRow + row, column=startCol + col).value = "%02.02f" % (s2profit / 100000000)
						s2valid = True
				elif col == COLOFFSET['Q3净利']:
					row_s3 = stockInfo.profit.loc[lambda df:df['REPORT_DATE'] =='%s 00:00:00' % reportDates[2],:]
					if len(row_s3):
						s3profit = row_s3['NETPROFIT'].values[0]
						ws.cell(row=startRow + row, column=startCol + col).value = "%02.02f" % (s3profit / 100000000)
						s3valid = True
				else:
					row_s4 = stockInfo.profit.loc[lambda df:df['REPORT_DATE'] =='%s 00:00:00' % reportDates[3],:]
					if len(row_s4):
						s4profit = row_s4['NETPROFIT'].values[0]
						ws.cell(row=startRow + row, column=startCol + col).value = "%02.02f" % (s4profit / 100000000)
						s4valid = True
				continue
		
		# print("row", row, ", ", s1valid, s2valid, s3valid, s4valid)
		if all((s1valid, s2valid, s3valid, s4valid)):
			resultProfit = sum((s1profit, s2profit, s3profit, s4profit)) / 100000000
			lastProfit = ws.cell(row=startRow + row - 1, column=startCol + COLOFFSET['净利润']).value
			if lastProfit:
				ws.cell(row=startRow + row, column=startCol + COLOFFSET['增长率']).value = ((resultProfit / float(lastProfit)) - 1)
			ws.cell(row=startRow + row, column=startCol + COLOFFSET['净利润']).value = "%02.02f" % resultProfit
		ws.cell(row=startRow + row, column=startCol + col).value = ""

	font = Font(bold=True)
	ws.cell(row=startRow, column=startCol).value = stockInfo.code
	ws.cell(row=startRow + 1, column=startCol).value = stockInfo.name
	ws.cell(row=startRow, column=startCol).font = font
	ws.cell(row=startRow + 1, column=startCol).font = font

	ws.cell(row=startRow, column=startCol + COLOFFSET['摘要']).value = "pe-ttm"
	ws.cell(row=startRow + 1, column=startCol + COLOFFSET['摘要']).value = stockInfo.pe_ttm

	ws.cell(row=startRow + 2, column=startCol + COLOFFSET['摘要']).value = "price"
	ws.cell(row=startRow + 3, column=startCol + COLOFFSET['摘要']).value = stockInfo.price

	# calc profit
	for rowIdx in (2, 3, 4, 5):
		ws.cell(row=startRow + rowIdx, column=startCol + COLOFFSET['净利润']).value = '=%s*(1+%s)' % (ws.cell(row=startRow + rowIdx - 1, column=startCol + COLOFFSET['净利润']).coordinate, ws.cell(row=startRow + rowIdx, column=startCol + COLOFFSET['增长率']).coordinate)

	# calc market value
	for rowIdx in (1, 2, 3, 4, 5):
		for colIdx in (COLOFFSET['低估估值'], COLOFFSET['合理估值'], COLOFFSET['高估估值']):
			ws.cell(row=startRow + rowIdx, column=startCol + colIdx).value = '=%s*%s' % (ws.cell(row=startRow + rowIdx, column=startCol + 3).coordinate, ws.cell(row=startRow, column=startCol + colIdx).coordinate)

		# calc quarter growth
		quarterCoordinate = []
		for quarterNum in (1, 2, 3, 4):
			beforeQuarter = ws.cell(row=startRow + rowIdx - 1, column=startCol + COLOFFSET['Q%d净利' % quarterNum])
			currentQuarter = ws.cell(row=startRow + rowIdx, column=startCol + COLOFFSET['Q%d净利' % quarterNum])
			ws.cell(row=startRow + rowIdx, column=startCol + COLOFFSET['Q%d增长' % quarterNum]).value = '=IFERROR((%s/%s)-1, "")' % (currentQuarter.coordinate, beforeQuarter.coordinate)
			quarterCoordinate.append(currentQuarter.coordinate)
		ws.cell(row=startRow + rowIdx, column=startCol + COLOFFSET['净利Sum']).value = '=IFERROR(' + '+'.join(quarterCoordinate) + ', "")'

	# current market value
	for i in range(6):
		ws.cell(row=startRow + i, column=startCol + COLOFFSET['当前市值']).value = stockInfo.market_value

	# setup style
	for idx, column in enumerate(ws.columns):
		if idx in (COLOFFSET['代码&名称'], COLOFFSET['摘要']):
			for cell in column:
				cell.fill = PatternFill('solid', fgColor='CCCCCC')
		elif idx in (COLOFFSET['增长率'],):
			for cell in column:
				cell.number_format = '0.00%'
		elif idx in (COLOFFSET['净利润'], COLOFFSET['低估估值'], COLOFFSET['合理估值'], COLOFFSET['高估估值'], COLOFFSET['当前市值']):
			for cell in column:
				cell.number_format = '0.00'
			if idx in (COLOFFSET['低估估值'], COLOFFSET['合理估值'], COLOFFSET['高估估值']):
				column[0].fill = PatternFill('solid', fgColor='EEEEEE')

	# add line chart
	if CHART:
		drawLineChart(ws, startRow, startCol)


def clearAllChart(worksheet):
    worksheet._charts = []


def drawLineChart(worksheet, startRow, startCol):
    c1 = LineChart()
    c1.title = worksheet.cell(row=startRow + 1, column=startCol + COLOFFSET['代码&名称']).value
    c1.style = 13

    data = Reference(worksheet, min_col=startCol + COLOFFSET['低估估值'], min_row=startRow + 1, max_col=COLOFFSET['当前市值'] + 1, max_row=startRow + 5)
    c1.add_data(data)

    for seriesIdx, color, lblTitle in zip(range(4), ("33CC00", "336699", "CC0000", "969696"), ("F1", "G1", "H1", "I1")):
        s = c1.series[seriesIdx]
        s.smooth = True
        s.graphicalProperties.line.solidFill = color
        s.title = SeriesLabel(StrRef(lblTitle))
        s.dLbls = DataLabelList()
        if seriesIdx == 3:
            d = DataLabel(4)
            d.showVal = True
            d.dLblPos = 'ctr'
            s.dLbls.dLbl.append(d)
        else:
            for t in range(5):
                d = DataLabel(t)
                d.showVal = True
                d.dLblPos = 'ctr'
                s.dLbls.dLbl.append(d)

    worksheet.add_chart(c1, worksheet.cell(row=startRow, column=startCol + COLOFFSET['当前市值'] + 1).coordinate)


if __name__ == "__main__":
	# generatePositions("position.xlsx")
	trackMarketValue("valueTrack.xlsx")
