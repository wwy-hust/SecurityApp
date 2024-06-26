# -*- coding:utf-8 -*-
import akshare as ak
import pandas as pd
import numpy as np
import math
import os, sys
import openpyxl
import traceback
import argparse

from openpyxl.styles import Alignment

from StockInfoLibrary.StockInfoClass import StockInfoProxy
from StockInfoLibrary.AkShareDataHelper import *
from StockInfoLibrary.TypeDefine import CodeType, TypeInPortfolio

from StockInfoLibrary.FutuAPIDataHelper import CloseQuoteContext

Stock = {}

# 总资产 & 净资产，用于计算仓位
TOTAL_MARKET_VALUE = 0
TOTAL_NET_MARKET_VALUE = 0

TOTAL_ASSET_IN_A = 0
TOTAL_ASSET_IN_HK = 0
TOTAL_ASSET_IN_US = 0

def getStockCodeDictFromExcel(excelPath, selectCols):
	assert len(selectCols) == 2
	if os.path.exists(excelPath):
		sheet = pd.read_excel(excelPath)
		codeAndNum = sheet[selectCols]
		cols = (codeAndNum[t] for t in selectCols)
		return dict(zip(*cols))


def generatePositions(positionFilePath):
	global Stock, TOTAL_MARKET_VALUE, TOTAL_NET_MARKET_VALUE
	global TOTAL_ASSET_IN_A, TOTAL_ASSET_IN_HK, TOTAL_ASSET_IN_US
	retDict = getStockCodeDictFromExcel(positionFilePath, ['Code', 'Num'])
	if retDict:
		Stock = retDict

	ret = []
	for stock, num in Stock.items():
		stockInfo = StockInfoProxy(stock)
		# try:
		stockInfo.fetchCodeData()
		# except:
		# 	print("Error")
		print(stockInfo)
		
		realValue = int(num * stockInfo.real_price)
		ret.append((stockInfo.originCode, stockInfo.typeInPorfolio, stockInfo.name, stockInfo.price, stockInfo.real_price, num, realValue, stockInfo.dividend_ratio_ttm))

		if stockInfo.code_type == CodeType.CURRENCY:
			TOTAL_NET_MARKET_VALUE += realValue
			
		else:
			TOTAL_MARKET_VALUE += realValue
			TOTAL_NET_MARKET_VALUE += realValue

			if stockInfo.typeInPorfolio == TypeInPortfolio.HK:
				TOTAL_ASSET_IN_HK += realValue
			elif stockInfo.typeInPorfolio == TypeInPortfolio.US:
				TOTAL_ASSET_IN_US += realValue
			elif stockInfo.typeInPorfolio == TypeInPortfolio.A:
				TOTAL_ASSET_IN_A += realValue


	NetTotalAsset = 0
	TotalAsset = 0
	for entry in ret:
		NetTotalAsset += entry[6]
		if entry[6] > 0:
			TotalAsset += entry[6]

	percentListInNetTotalAsset = []
	percentListInTotalAsset = []
	for entry in ret:
		percent = float('%.4f' % float(entry[6] / NetTotalAsset))
		percentListInNetTotalAsset.append(percent)
		percent1 = float('%.4f' % float(entry[6] / TotalAsset))
		percentListInTotalAsset.append(percent1)

	df = pd.DataFrame(ret, columns=['Code', 'Type', 'Name', 'Price', 'RealPrice', 'Num', 'Value', 'Dividend'])
	df['PercentInNetTotalAsset'] = percentListInNetTotalAsset
	df['PercentInTotalAsset'] = percentListInTotalAsset
	df.sort_values(by='Value', ascending=False, inplace=True)
	print(df)
	# df.to_excel(positionFilePath, index=False)

	AverageRatioStr = "="
	wb = openpyxl.load_workbook(positionFilePath)
	ws = wb.active
	for rowIdx, row in enumerate(df.iterrows(), 2):
		ws.cell(row=rowIdx, column=1).value = row[1]['Code']

		ws.cell(row=rowIdx, column=2).value = row[1]['Name']
		ws.cell(row=rowIdx, column=2).alignment = Alignment(horizontal='center', vertical='center')

		ws.cell(row=rowIdx, column=3).value = row[1]['Price']
		ws.cell(row=rowIdx, column=3).number_format = '0.000'

		ws.cell(row=rowIdx, column=4).value = row[1]['RealPrice']
		ws.cell(row=rowIdx, column=4).number_format = '0.000'

		ws.cell(row=rowIdx, column=5).value = row[1]['Num']

		ws.cell(row=rowIdx, column=6).value = '=%s*%s' % (ws.cell(row=rowIdx, column=4).coordinate, ws.cell(row=rowIdx, column=5).coordinate)
		ws.cell(row=rowIdx, column=6).number_format = "0"

		ws.cell(row=rowIdx, column=7).value = round(row[1]['Dividend'] / 100, 4)
		ws.cell(row=rowIdx, column=7).number_format = "0.00%"


		ws.cell(row=rowIdx, column=8).value = '=%s/K1' % (ws.cell(row=rowIdx, column=6).coordinate)
		ws.cell(row=rowIdx, column=8).alignment = Alignment(horizontal='center', vertical='center')
		ws.cell(row=rowIdx, column=8).number_format = '0.00%'

		ws.cell(row=rowIdx, column=9).value = '=%s/N1' % (ws.cell(row=rowIdx, column=6).coordinate)
		ws.cell(row=rowIdx, column=9).alignment = Alignment(horizontal='center', vertical='center')
		ws.cell(row=rowIdx, column=9).number_format = '0.00%'

		AverageRatioStr += "%s*%s+" % (ws.cell(row=rowIdx, column=7).coordinate, ws.cell(row=rowIdx, column=8).coordinate)

	AverageRatioStr += '0'
	ws['K1'].value = '=SUM(F2:F%d)' % (rowIdx)	# NetTotalAsset
	ws['N1'].value = '%d' % TotalAsset	# TotalAsset
	ws['L1'].value = '=%02f' % (float(TOTAL_MARKET_VALUE) / float(TOTAL_NET_MARKET_VALUE))	# 杠杆率
	ws['L2'].value = '=SUM(G2:G4)'
	ws['L3'].value = '=SUM(G2:G6)'
	ws['L4'].value = '=SUM(G2:G11)'

	ws['N2'].value = '=SUM(H2:H4)'
	ws['N3'].value = '=SUM(H2:H6)'
	ws['N4'].value = '=SUM(H2:H11)'

	ws['P1'].value = AverageRatioStr
	ws['P1'].number_format = '0.00%'

	wb.save(positionFilePath)

	print("NetTotalAsset:%d | TotalAsset:%d | POSITION:%02f" % (NetTotalAsset, TotalAsset, (float(TOTAL_MARKET_VALUE) / float(TOTAL_NET_MARKET_VALUE))))
	print("A:%.2f | HK:%.2f | US:%.2f" % ((float)(TOTAL_ASSET_IN_A)/(float)(TotalAsset), (float)(TOTAL_ASSET_IN_HK)/(float)(TotalAsset),(float)(TOTAL_ASSET_IN_US)/(float)(TotalAsset)))
	
	CloseQuoteContext()


if __name__ == "__main__":
	parser = argparse.ArgumentParser(description='这是一个示例程序。')
	parser.add_argument('--update', dest='doUpdateAkshareData', action='store_const',
                    const=True, default=False,
                    help='update all akshare cache')
	args = parser.parse_args()

	if args.doUpdateAkshareData:
		CacheAllAKShareData()
	else:
		generatePositions("position.xlsx")
	# outputFinancialInfo("valueTrack.xlsx")
